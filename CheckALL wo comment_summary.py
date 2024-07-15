# -*- coding: utf-8 -*-
"""
Created on Sun Feb 18 00:05:21 2024

@author: JIA
"""

import subprocess
import time
import os
import datetime
from typing import Tuple

import win32ui
import dde
import pandas as pd
import requests
import talib
from openpyxl import load_workbook
from talib import abstract

# Constants
WINDOW_WEEK = 5
WINDOW_MONTH = 21
WINDOW_SEASON = 62
WINDOW_YEAR = 248

V_PER_05S = 4.5 * 3600 / 5          # per 5 secs
V_PER_10S = 4.5 * 3600 / 10         # per 10 secs
V_PER_15S = 4.5 * 3600 / 15         # per 15 secs

# LINE Notify token
TOKEN = 'nkss6NkZp619xiVjXtQo0psTWIrgimB1PT5gbQGl8Gh'

# Debug mode
DEBUG = True



def send_line_notify(message: str) -> None:
    """Send LINE Notify message."""
    headers = {"Authorization": "Bearer " + TOKEN}
    data = {'message': message}
    requests.post('https://notify-api.line.me/api/notify', headers=headers, data=data)



def calculate_rsv(df: pd.DataFrame, n: int = 9) -> pd.Series:
    """Calculate Relative Strength Value (RSV)."""
    return ((df['Close'] - df['Low'].rolling(window=n).min()) /
            (df['High'].rolling(window=n).max() - df['Low'].rolling(window=n).min())) * 100

def calculate_k(df: pd.DataFrame, rsv: pd.Series, n: int = 9) -> pd.Series:
    """Calculate K value."""
    k = [None] * len(df)
    k[n - 2] = 50.  # Set the first value of K to 50
    for i in range(n - 1, len(df)):
        k[i] = (2. * k[i - 1] + rsv[i]) / 3.
    return k

def calculate_d(df: pd.DataFrame, k: pd.Series, n: int = 9) -> pd.Series:
    """Calculate D value."""
    d = [None] * len(df)
    d[n - 2] = 50.  # Set the first value of D to 50
    for i in range(n - 1, len(df)):
        d[i] = (2. * d[i - 1] + k[i]) / 3.
    return d

def get_stock_data(conversation, symbol: str) -> Tuple[float, ...]:
    """Return stock data ('Open', 'High', 'Low', 'Close', 'Volume', 'K', 'D', 'BBAND_u', 'BBAND_l', DataFrame)."""
    window = WINDOW_YEAR + 20
    
    try:
        requested_data = conversation.Request(f"{symbol}.TW-Day-{window}")
        rows = requested_data.split(';')
    except Exception as e:
        print(f"REQUEST FAILED ON STOCK SYMBOL {symbol}: {e}")
        return (None, None, None, None, None, None, None, None, None, None)
        

    if len(rows) < window:
        return (None,) * 10

    data = [row.split(',') for row in rows]
    df = pd.DataFrame(data)
    df.columns = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume']
    
    list_float = ['Open', 'High', 'Low', 'Close', 'Volume']
    df[list_float] = df[list_float].astype(float)

    rsv = calculate_rsv(df)
    df['K'] = calculate_k(df, rsv)
    df['D'] = calculate_d(df, df['K'])

    df['BBAND_u'], df['BBAND_m'], df['BBAND_l'] = talib.abstract.BBANDS(df['Close'], timeperiod=21, nbdevup=2.0,
                                                                         nbdevdn=2.0, matype=0)

    o = float(df.iloc[-1]['Open'])
    h = float(df.iloc[-1]['High'])
    l = float(df.iloc[-1]['Low'])
    c = float(df.iloc[-1]['Close'])
    v = float(df.iloc[-1]['Volume'])
    k = float(df.iloc[-1]['K'])
    d = float(df.iloc[-1]['D'])
    bu = float(df.iloc[-1]['BBAND_u'])
    bm = float(df.iloc[-1]['BBAND_m'])
    bl = float(df.iloc[-1]['BBAND_l'])
    return (o, h, l, c, v, k, d, bu, bl, df)



def check_today(symbol: str, name: str, industry: str, 
                conversation, summary: pd.DataFrame, 
                ht: float , lt: float ,
                must_check: bool , 
                file_quit: str , file_today: str ,
                white_list_of_symbol, black_list_of_symbol):
    """Check stock data for today and send notifications if necessary."""
    o, h, l, c, v, k, d, bu, bl, df = get_stock_data(conversation, symbol)

    if o is None and h is None and l is None and c is None:
        if DEBUG:
            with open(file_quit, 'a', encoding='utf-8') as f:
                f.write(f"Symbol: {symbol}, Name: {name}. Data from the server is shorter than 1 year.\n")
        return False, None

    if must_check:
        if v < V_PER_15S:
            if DEBUG:
                with open(file_quit, 'a', encoding='utf-8') as f:
                    f.write(f"Symbol: {symbol}, Name: {name}, Volume: {v}. Too Few Volume, period longer than 15 sec.\n")
            return False, None
    else:
        if v < V_PER_05S:
            if DEBUG:
                with open(file_quit, 'a', encoding='utf-8') as f:
                    f.write(f"Symbol: {symbol}, Name: {name}, Volume: {v}. Too Few Volume, period longer than 5 sec.\n")
            return False, None

        high_season = df['High'].rolling(window=WINDOW_SEASON).max().iloc[-2]
        low_season = df['Low'].rolling(window=WINDOW_SEASON).min().iloc[-2]
        if high_season / low_season < 1.10:
            if DEBUG:
                with open(file_quit, 'a', encoding='utf-8') as f:
                    f.write(f"Symbol: {symbol}, Name: {name}. Price Variation Too Low.\n")
                    f.write(f"過去3個月最高價格: {high_season}\n")
                    f.write(f"過去3個月最低價格: {low_season}\n")
                    f.write(f"過去3個月最高/最低價格比: {high_season / low_season:.2f}\n")
            return False, None

    message = (
        f"Symbol:\t{symbol}\n"
        f"Name:\t{name}\n"
        f"Industry:\t{industry}\n\n"
    )
    notify = False

    if symbol == "TSE":
        notify = True

    if must_check and ht != 0 and c >= ht:
        message += (
            "價格接近高點:\n"
            f"價格: {c}\n"
            f"高點: {ht:.2f}\n\n"
        )
        notify = True

    if lt != 0 and c <= lt:
        message += (
            "價格接近低點:\n"
            f"價格: {c}\n"
            f"低點: {lt:.2f}\n\n"
        )
        notify = True

    if l < 1.01 * bl:
        message += (
            "價格接近布林下軌:\n"
            f"最低價: {l}\n"
            f"布林下軌: {bl:.2f}\n\n"
        )
        notify = True

    if 15. <= k <= 25.:
        message += (
            "K值接近20:\n"
            f"K值: {k:.2f}\n\n"
        )
        notify = True

    if must_check and 75. <= k <= 85.:
        message += (
            "K值接近80:\n"
            f"K值: {k:.2f}\n\n"
        )
        notify = True

    if notify:
        message += "\n\n"
        if DEBUG:
            with open(file_today, 'a', encoding='utf-8') as f:
                f.write(message)
        else:
            send_line_notify(message)
            
        return True, summary_per_symbol(symbol, name, industry, df, summary, 
                                        white_list_of_symbol, black_list_of_symbol)
    else:
        return True, None


def check_history(symbol: str, name: str, industry: str, conversation,
                  file_history: str) -> None:
    """Check historical stock data and send notifications if necessary."""
    o, h, l, c, v, k, d, bu, bl, df = get_stock_data(conversation, symbol)

    message = (
        f"Symbol:\t{symbol}\n"
        f"Name:\t{name}\n"
        f"Industry:\t{industry}\n\n"
    )
    notify = False

    c_yesterday = df.iloc[-2]['Close']
    volume_month = df['Volume'].rolling(window=WINDOW_MONTH).mean().iloc[-2]
    if c >= c_yesterday * 1.05 and v >= volume_month * 1.5:
        message += (
            "價格、成交量均上漲:\n"
            f"今日價格: {c}\n"
            f"昨日價格: {c_yesterday}\n"
            f"今日/昨日價格比: {c / c_yesterday:.2f}\n"
            f"今日成交量: {v}\n"
            f"過去1個月平均成交量: {volume_month:.2f}\n"
            f"今日/1個月平均成交量比: {v / volume_month:.2f}\n\n"
        )
        notify = True

    low_season = df['Low'].rolling(window=WINDOW_SEASON).min().iloc[-2]
    if l / low_season <= 1 + 0.05:
        message += (
            "今日最低價接近過去3個月的最低價:\n"
            f"今日最低價: {l}\n"
            f"過去3個月最低價: {low_season}\n"
            f"今日最低價 / 過去3個月最低價 : {l / low_season:.2f}\n\n"
        )
        notify = True

    high_season = df['High'].rolling(window=WINDOW_SEASON).max().iloc[-2]
    if h / high_season >= 1 - 0.05:
        message += (
            "今日最高價接近過去3個月的最高價:\n"
            f"今日最高價: {h}\n"
            f"過去3個月最高價: {high_season}\n"
            f"今日最高價 / 過去3個月最高價 : {h / high_season:.2f}\n\n"
        )
        notify = True

    if notify:
        message += "\n\n"
        if DEBUG:
            with open(file_history, 'a', encoding='utf-8') as f:
                f.write(message)
        else:
            send_line_notify(message)



def summary_per_symbol(symbol: str, name: str, industry: str, 
                       df: pd.DataFrame, summary: pd.DataFrame, 
                       white_list_of_symbol, black_list_of_symbol) -> pd.DataFrame:
    Open_m3 = df.iloc[-3]['Open']
    Open_m2 = df.iloc[-2]['Open']
    Open_m1 = df.iloc[-1]['Open']
    
    High_m3 = df.iloc[-3]['High']
    High_m2 = df.iloc[-2]['High']
    High_m1 = df.iloc[-1]['High']
    
    Low_m3 = df.iloc[-3]['Low']
    Low_m2 = df.iloc[-2]['Low']
    Low_m1 = df.iloc[-1]['Low']
    
    Close_m3 = df.iloc[-3]['Close']
    Close_m2 = df.iloc[-2]['Close']
    Close_m1 = df.iloc[-1]['Close']
    Gain_m3 = df.iloc[-3]['Close'] - df.iloc[-4]['Close']
    Gain_m2 = df.iloc[-2]['Close'] - df.iloc[-3]['Close']
    Gain_m1 = df.iloc[-1]['Close'] - df.iloc[-2]['Close']
    
    High_Season = df['High'].rolling(window=WINDOW_SEASON).max().iloc[-2]
    Approx_High_Season = (High_m1 / High_Season) >= (1. - 0.05)
    GTE_High_Season = High_m1 >= High_Season
    
    Low_Season = df['Low'].rolling(window=WINDOW_SEASON).min().iloc[-2]
    Approx_Low_Season = (Low_m1 / Low_Season) <= (1. + 0.05)
    LTE_Low_Season = Low_m1 <= Low_Season
    
    High_Div_Low_Season = (High_Season / Low_Season) - 1.
    
    Vol_m3 = df.iloc[-3]['Volume']
    Vol_m2 = df.iloc[-2]['Volume']
    Vol_m1 = df.iloc[-1]['Volume']
    
    Vol_LT_1080 = Vol_m1 < 1080.
    Vol_BTW_1080_1620 = (1080. <= Vol_m1) and (Vol_m1 < 1620.)
    Vol_BTW_1620_3240 = (1620. <= Vol_m1) and (Vol_m1 < 3240.)
    Vol_GTE_3240 = 3240. <= Vol_m1
    
    Vol_Mean_Month = df['Volume'].rolling(window=WINDOW_MONTH).mean().iloc[-2]
    Vol_Div_Vol_Mean_Month = Vol_m1 / Vol_Mean_Month
    
    K_m3 = df.iloc[-3]['K']
    K_m2 = df.iloc[-2]['K']
    K_m1 = df.iloc[-1]['K']
    
    K_BTW_15_25 = (15. <= K_m1) and (K_m1 < 25.)
    K_LTE_20 = K_m1 <= 20.
    K_BTW_75_85 = (75. <= K_m1) and (K_m1 < 85.)
    K_GTE_80 = 80. <= K_m1
    
    D_m3 = df.iloc[-3]['D']
    D_m2 = df.iloc[-2]['D']
    D_m1 = df.iloc[-1]['D']
    
    B_Band_U_m3 = df.iloc[-3]['BBAND_u']
    B_Band_U_m2 = df.iloc[-2]['BBAND_u']
    B_Band_U_m1 = df.iloc[-1]['BBAND_u']
    Approx_B_Band_U = ((1. - 0.01) * B_Band_U_m1) < High_m1
    
    B_Band_L_m3 = df.iloc[-3]['BBAND_l']
    B_Band_L_m2 = df.iloc[-2]['BBAND_l']
    B_Band_L_m1 = df.iloc[-1]['BBAND_l']
    Approx_B_Band_L = Low_m1 < ((1. + 0.01) * B_Band_L_m1)
    
    # white lists
    ETF0050 = False
    ETF0056 = False
    ETF00878 = False
    ETF00921 = False
    TwMid100 = False
    ROE_Season_GT_5 = False
    Warrant = False
    Future = False
    Option = False
    High_Yield = False
    Ind_Leader = False
    
    for white_list_name in white_list_of_symbol:
        match white_list_name:
            case 'ETF成分股0050':
                ETF0050 = True
            case 'ETF成分股0056':
                ETF0056 = True
            case 'ETF成分股00878':
                ETF00878 = True
            case 'ETF成分股00921':
                ETF00921 = True
            case '台灣中型100':
                TwMid100 = True
            case '本季度ROE 5%以上':
                ROE_Season_GT_5 = True
            case '權證標的_上市':
                Warrant = True
            case '股票期貨標的_上市':
                Future = True
            case '股票選擇權標的_上市':
                Option = True
            case '高殖利率_權證':
                High_Yield = True
            case '產業龍頭':
                Ind_Leader = True
    
    # black lists
    ROE_3Year_LT_5 = False
    Full_Cash = False
    Vol_Year_LT_1620 = False
    
    for black_list_name in black_list_of_symbol:
        match black_list_name:
            case 'ROE連3年0~5':
                ROE_3Year_LT_5 = True
            case '全額交割股':
                Full_Cash = True
            case '成交量去年全年日均量0~1620':
                Vol_Year_LT_1620 = True
        
    # Create a new row of data as list
    new_row = [symbol, name, industry, 
               Open_m3, Open_m2, Open_m1, 
               High_m3, High_m2, High_m1, 
               Low_m3, Low_m2, Low_m1, 
               Close_m3, Gain_m3, Close_m2, Gain_m2, Close_m1, Gain_m1, 
               High_Season, Approx_High_Season, GTE_High_Season, 
               Low_Season, Approx_Low_Season, LTE_Low_Season, 
               High_Div_Low_Season, 
               Vol_m3, Vol_m2, Vol_m1, 
               Vol_LT_1080, Vol_BTW_1080_1620, Vol_BTW_1620_3240, Vol_GTE_3240, 
               Vol_Mean_Month, Vol_Div_Vol_Mean_Month, 
               K_m3, K_m2, K_m1, 
               K_BTW_15_25, K_LTE_20, K_BTW_75_85, K_GTE_80, 
               D_m3, D_m2, D_m1, 
               B_Band_U_m3, B_Band_U_m2, B_Band_U_m1, Approx_B_Band_U, 
               B_Band_L_m3, B_Band_L_m2, B_Band_L_m1, Approx_B_Band_L, 
               ETF0050, ETF0056, ETF00878, ETF00921, 
               TwMid100, 
               ROE_Season_GT_5, 
               Warrant, Future, Option, 
               High_Yield, Ind_Leader, 
               ROE_3Year_LT_5, Full_Cash, Vol_Year_LT_1620]
    
    # Convert the list to a DataFrame
    df_row = pd.DataFrame([new_row], columns=summary.columns)
    
    return df_row
    



def generate_list(parent_dir: str) -> list:
    """Generate a list of all sub-folders and dictionary of the last file in each sub-folder."""
    returned_list = []

    for subdir, _, files in os.walk(parent_dir):
        subfolder_name = os.path.basename(subdir)

        if files:
            sorted_files = sorted(files)
            last_file = sorted_files[-1]
            last_file_path = os.path.join(subdir, last_file)

            try:
                if last_file_path.endswith('.csv'):
                    df = pd.read_csv(last_file_path)
                elif last_file_path.endswith('.xlsx'):
                    df = pd.read_excel(last_file_path)
                else:
                    continue

                if not df.empty:
                    df.iloc[:, 0] = df.iloc[:, 0].astype("string").str.replace('=', '').str.replace('"', '')
                    column_dict = dict(zip(df.iloc[:, 0], [1] * len(df)))
                    returned_list.append((subfolder_name, column_dict))
            except Exception as e:
                print(f"Error reading file {last_file_path}: {e}")

    return returned_list



def main():
    """Main function.
    Launch and Create a XQLite DDE server, 
    Fetch and Calculate price and volume data of given stock, 
    """
    # Launch XQLite
    program_path = r'C:\SysJust\XQLite\daqxqlite.exe'
    xqlite = subprocess.Popen(program_path)
    time.sleep(60)  # Wait for XQLite to open

    # Create a DDE server and conversation
    server = dde.CreateServer()
    server.Create("TestClient")
    conversation = dde.CreateConversation(server)
    conversation.ConnectTo("XQLITE", "Kline")



    # Generate lists
    white_list = generate_list('D:\\DJC\\股票\\白名單')
    black_list = generate_list('D:\\DJC\\股票\\黑名單')
    other_list = generate_list('D:\\DJC\\股票\\其他名單')
    must_check_list = ["", ""]



    # Debug
    today = datetime.date.today()
    formatted_date = today.strftime("%Y-%m-%d")

    file_quit_name = f'quit message_{formatted_date}.txt'
    file_today_name = f'today message_{formatted_date}.txt'
    file_history_name = f'history message_{formatted_date}.txt'
    file_list_name = f'list message_{formatted_date}.txt'

    if DEBUG:
        file_quit = open(file_quit_name, 'w', encoding='utf-8')
        file_today = open(file_today_name, 'w', encoding='utf-8')
        file_history = open(file_history_name, 'w', encoding='utf-8')
        file_list = open(file_list_name, 'w', encoding='utf-8')



    # Summary 
    # Column title:
    # Symbol	Name	Industry	
    # "Open, 2日前"	"Open, 1日前"	"Open, 今日"	
    # "High, 2日前"	"High, 1日前"	"High, 今日"	
    # "Low, 2日前"	"Low, 1日前"	"Low, 今日"	
    # "Close, 2日前"	"漲幅, 2日前"	"Close, 1日前"	"漲幅, 1日前"	"Close, 今日"	"漲幅, 今日"	
    # "過去3個月最高價"	"接近過去最高價"	"大於過去最高價"	
    # "過去3個月最低價"	"接近過去最低價"	"小於過去最低價"	
    # "過去3個月高/低價差比"	
    # "Volume(張), 2日前"	"Volume(張), 1日前"	"Volume(張), 今日"	
    # Vol< 1080	1080< Vol< 1620	1620< Vol< 3240	3240< Vol	
    # "過去1個月均量(張)"	"今日量/過去均量(張)"	
    # "K, 2日前"	"K, 1日前"	"K, 今日"	
    # 15≤ K≤ 25	K≤ 20	75≤ K≤ 85	80≤ K	
    # "D, 2日前"	"D, 1日前"	"D, 今日"	
    # "布林上, 2日前"	"布林上, 1日前"	"布林上, 今日"	"接近布林上"	
    # "布林下, 2日前"	"布林下, 1日前"	"布林下, 今日"	"接近布林下"	
    # ETF成分股 0050	ETF成分股 0056	ETF成分股 00878	ETF成分股 00921	
    # 臺灣中型100指數成分股
    # 本季度ROE 5%以上	
    # 權證標的 上市	股票期貨標的 上市	股票選擇權標的 上市	
    # 高殖利率 權證	產業龍頭	
    # ROE連3年 5以下	全額交割股(full-cash delivery stock)	成交量去年全年日均量 0~1620

    list_title = ["Symbol", "Name", "Industry", 
                  "Open_m3", "Open_m2", "Open_m1", 
                  "High_m3", "High_m2", "High_m1", 
                  "Low_m3", "Low_m2", "Low_m1", 
                  "Close_m3", "Gain_m3", "Close_m2", "Gain_m2", "Close_m1", "Gain_m1", 
                  "High_Season", "Approx_High_Season", "GTE_High_Season", 
                  "Low_Season", "Approx_Low_Season", "LTE_Low_Season", 
                  "High_Div_Low_Season", 
                  "Vol_m3", "Vol_m2", "Vol_m1", 
                  "Vol_LT_1080", "Vol_BTW_1080_1620", "Vol_BTW_1620_3240", "Vol_GTE_3240", 
                  "Vol_Mean_Month", "Vol_Div_Vol_Mean_Month", 
                  "K_m3", "K_m2", "K_m1", 
                  "K_BTW_15_25", "K_LTE_20", "K_BTW_75_85", "K_GTE_80", 
                  "D_m3", "D_m2", "D_m1", 
                  "B_Band_U_m3", "B_Band_U_m2", "B_Band_U_m1", "Approx_B_Band_U", 
                  "B_Band_L_m3", "B_Band_L_m2", "B_Band_L_m1", "Approx_B_Band_L",
                  "ETF0050", "ETF0056", "ETF00878", "ETF00921", 
                  "TwMid100", 
                  "ROE_Season_GT_5", 
                  "Warrant", "Future", "Option", 
                  "High_Yield", "Ind_Leader", 
                  "ROE_3Year_LT_5", "Full_Cash", "Vol_Year_LT_1620"]
    
    # Create a DataFrame
    summary = pd.DataFrame(columns = list_title)
    
    
    
    # 加權指數
    symbol, name, industry = "TSE", "加權指數", "-"
    print(f'symbol: {symbol}, name: {name}, industry: {industry}')
    
    must_check = True
    white_list_of_symbol = []
    black_list_of_symbol = []
    ht, lt = 0, 0
    
    result, summary_symbol = check_today(symbol, name, industry, 
                                         conversation, summary, 
                                         ht, lt, must_check,
                                         file_quit_name, file_today_name, 
                                         white_list_of_symbol, black_list_of_symbol)

    summary = summary_symbol
    
    check_history(symbol, name, industry, conversation, file_history_name)



    # Load stock list
    file = r'D:\DJC\股票\上市清單\上市清單.xlsx'
    wb_obj = load_workbook(filename=file, data_only=True)
    wsheet = wb_obj.worksheets[0]

    # Loop through each stock in stock list
    for row in wsheet.iter_rows(min_row=2, values_only=True):
        if str(row[0])[0].isdigit() and row[1] is not None:
            
            if row[0][-1].isalpha() and row[0][-1] != "U":
                continue

            symbol, name, industry = row[0], row[1], row[6]

            if not symbol:
                continue

            if DEBUG:
                with open(file_list_name, 'a', encoding='utf-8') as f:
                    f.write(', '.join([symbol, name, industry]) + "\n")

            # Check white list
            must_check = False
            white_list_of_symbol = []
            for white_list_name, white_list_dict in white_list:
                if symbol in white_list_dict:
                    must_check = True
                    white_list_of_symbol.append(white_list_name)

            if white_list_of_symbol and DEBUG:
                with open(file_list_name, 'a', encoding='utf-8') as f:
                    f.write(f'Symbol: {symbol}, Name: {name}, Industry: {industry} is in white list:\n')
                    f.write(', '.join(white_list_of_symbol) + "\n\n")

            # Check black list
            black_list_of_symbol = []
            for black_list_name, black_list_dict in black_list:
                if symbol in black_list_dict:
                    black_list_of_symbol.append(black_list_name)

            if black_list_of_symbol and DEBUG:
                with open(file_list_name, 'a', encoding='utf-8') as f:
                    f.write(f'Symbol: {symbol}, Name: {name}, Industry: {industry} is in black list:\n')
                    f.write(', '.join(black_list_of_symbol) + "\n\n")
                
            if (black_list_of_symbol) and (not must_check):
                    continue

            # Check other list
            other_list_of_symbol = []
            for other_list_name, other_list_dict in other_list:
                if symbol in other_list_dict:
                    other_list_of_symbol.append(other_list_name)

            if other_list_of_symbol and DEBUG:
                with open(file_list_name, 'a', encoding='utf-8') as f:
                    f.write(f'Symbol: {symbol}, Name: {name}, Industry: {industry} is in other list:\n')
                    f.write(', '.join(other_list_of_symbol) + "\n\n")

            # Check stock data
            print(f'symbol: {symbol}, name: {name}, industry: {industry}')
            
            ht, lt = 0, 0
            
            result, summary_symbol = check_today(symbol, name, industry, 
                                                 conversation, summary, 
                                                 ht, lt, must_check,
                                                 file_quit_name, file_today_name, 
                                                 white_list_of_symbol, black_list_of_symbol)
            
            if summary_symbol is not None:
                # Concatenate DataFrame df_row to DataFrame summary
                if summary.empty:
                    summary = summary_symbol
                else:
                    summary = pd.concat([summary, summary_symbol], ignore_index=True)
                    
                # # Write white, black, and other lists to DataFrame summary
                # if not summary.empty:
                #     if summary.iloc[-1]['Symbol'] == symbol:
                #         # white lists
                #         for white_list_name in white_list_of_symbol:
                #             match white_list_name:
                #                 case 'ETF成分股0050':
                #                     summary.at[-1, 'ETF0050'] = True
                #                 case 'ETF成分股0056':
                #                     summary.at[-1, 'ETF0056'] = True
                #                 case 'ETF成分股00878':
                #                     summary.at[-1, 'ETF00878'] = True
                #                 case 'ETF成分股00921':
                #                     summary.at[-1, 'ETF00921'] = True
                #                 case '本季度ROE 5%以上':
                #                     summary.iloc[-1]['ROE_Season_GT_5'] = True
                #                 case '權證標的_上市':
                #                     summary.iloc[-1]['Warrant'] = True
                #                 case '股票期貨標的_上市':
                #                     summary.iloc[-1]['Future'] = True
                #                 case '股票選擇權標的_上市':
                #                     summary.iloc[-1]['Option'] = True
                #                 case '高殖利率_權證':
                #                     summary.iloc[-1]['High_Yield'] = True
                #                 case '產業龍頭':
                #                     summary.iloc[-1]['Ind_Leader'] = True

            if result:
                check_history(symbol, name, industry, conversation, file_history_name)
                


    # Close files and XQLite
    if DEBUG:
        file_quit.close()
        file_today.close()
        file_history.close()
        file_list.close()

    xqlite.terminate()
    time.sleep(10)



    # Summary 
    # Column data type
    list_float = ["Open_m3", "Open_m2", "Open_m1", 
                  "High_m3", "High_m2", "High_m1", 
                  "Low_m3", "Low_m2", "Low_m1", 
                  "Close_m3", "Gain_m3", "Close_m2", "Gain_m2", "Close_m1", "Gain_m1", 
                  "High_Season", "Low_Season", "High_Div_Low_Season", 
                  "Vol_m3", "Vol_m2", "Vol_m1", 
                  "Vol_Mean_Month", "Vol_Div_Vol_Mean_Month", 
                  "K_m3", "K_m2", "K_m1", 
                  "D_m3", "D_m2", "D_m1", 
                  "B_Band_U_m3", "B_Band_U_m2", "B_Band_U_m1", 
                  "B_Band_L_m3", "B_Band_L_m2", "B_Band_L_m1"]
    summary[list_float] = summary[list_float].astype(float)
    
    list_bool = ["Approx_High_Season", "GTE_High_Season", 
                 "Approx_Low_Season", "LTE_Low_Season", 
                 "Vol_LT_1080", "Vol_BTW_1080_1620", "Vol_BTW_1620_3240", "Vol_GTE_3240", 
                 "K_BTW_15_25", "K_LTE_20", "K_BTW_75_85", "K_GTE_80", 
                 "Approx_B_Band_U", "Approx_B_Band_L", 
                 "ETF0050", "ETF0056", "ETF00878", "ETF00921", 
                 "TwMid100",
                 "ROE_Season_GT_5", 
                 "Warrant", "Future", "Option", 
                 "High_Yield", "Ind_Leader", 
                 "ROE_3Year_LT_5", "Full_Cash", "Vol_Year_LT_1620"]
    summary[list_bool] = summary[list_bool].astype(bool)
    
    # Write summary to an Excel file
    summary.to_excel(f'summary_{formatted_date}.xlsx', index=False)



if __name__ == "__main__":
    main()